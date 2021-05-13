from os import walk,mkdir
from shutil import copy, move
from openpyxl import load_workbook


#加载excel

wb = load_workbook('staffinfo.xlsx')
ws = wb['Sheet1']

#注意这里 ws.rows 返回一个生成器对象，但是我们取值需要从第二行开始取
#那么我们可以用推导列表吧对象取出来放在list里，然后进行切片取值
ws = [x for x in ws.rows]

for x in ws[1:]:
    print(x[0].value,x[1].value)
    mkdir('company_staff/'+x[0].value)







